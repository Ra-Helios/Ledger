import openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
from .expense import get_expenses, project_stats
from .storage import load_project

EXPORT_DIR = os.path.join(os.path.dirname(__file__), '..', 'exports')
PALETTE = ["3B82F6","8B5CF6","10B981","F59E0B","EF4444","06B6D4","EC4899","84CC16","F97316"]

def _h(cell, bg="1e293b", fg="e2e8f0", bold=True, size=11, center=True):
    cell.font = Font(bold=bold, color=fg, size=size, name='Calibri')
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)

def _border():
    s = Side(style='thin', color="334155")
    return Border(left=s, right=s, top=s, bottom=s)


def export_project(slug):
    os.makedirs(EXPORT_DIR, exist_ok=True)
    proj = load_project(slug)
    expenses = get_expenses(slug)
    stats = project_stats(slug, expenses)
    cur = proj.get('currency') or '₹'
    proj_name = proj['name']
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Expenses ──────────────────────────────
    ws = wb.active
    ws.title = "Expenses"
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:J1')
    t = ws['A1']
    t.value = f"{proj['icon']} {proj_name}  —  Expense Report  |  {datetime.now().strftime('%d %b %Y')}"
    t.font = Font(bold=True, size=13, color="F8FAFC", name='Calibri')
    t.fill = PatternFill("solid", fgColor="0F172A")
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    hdrs = ["#", "Date", "Category", "Vendor", "Description", f"Amount ({cur})", "Mode", "Tags", "Notes", "ID"]
    widths = [5, 13, 22, 22, 35, 16, 14, 20, 25, 6]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        _h(c, bg="1E293B")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    mode_fill = {"Cash":"D1FAE5","Gpay":"DBEAFE","Bank Transfer":"EDE9FE","UPI":"FEF3C7","Cheque":"FCE7F3"}
    for ri, e in enumerate(expenses, 3):
        bg = "F8FAFC" if ri % 2 == 0 else "F1F5F9"
        vals = [ri-2, e['date'], e['category'], e['vendor'], e['description'],
                e['amount'], e['mode'], ', '.join(e.get('tags',[])), e.get('notes',''), e['id']]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = _border()
            cell.alignment = Alignment(vertical='center', wrap_text=(ci in [5,8,9]))
            if ci == 6:
                cell.number_format = f'#,##0.00'
                cell.font = Font(bold=True, color="1E3A5F", name='Calibri')
                cell.fill = PatternFill("solid", fgColor="DBEAFE")
            elif ci == 7:
                cell.fill = PatternFill("solid", fgColor=mode_fill.get(str(v), bg))
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.fill = PatternFill("solid", fgColor=bg)

    tr = len(expenses) + 3
    ws.merge_cells(f'A{tr}:E{tr}')
    tc = ws[f'A{tr}']
    tc.value = f"TOTAL  ({len(expenses)} entries)"
    _h(tc, bg="0F172A", center=False)
    tc.alignment = Alignment(horizontal='right', vertical='center')
    tot = ws.cell(row=tr, column=6, value=stats['total'])
    tot.number_format = '#,##0.00'
    _h(tot, bg="2563EB")
    ws.freeze_panes = 'A3'

    # ── Sheet 2: Category ──────────────────────────────────
    ws2 = wb.create_sheet("By Category")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells('A1:C1')
    _h(ws2['A1'], bg="0F172A"); ws2['A1'].value = "Category Breakdown"
    ws2.row_dimensions[1].height = 26
    for ci, h in enumerate(["Category","Total","% Share"],1):
        c = ws2.cell(row=2,column=ci,value=h); _h(c, bg="1E293B")
        ws2.column_dimensions[get_column_letter(ci)].width = [28,18,12][ci-1]
    gt = stats['total'] or 1
    for ri,(cat,amt) in enumerate(stats['category_breakdown'].items(),3):
        bg = PALETTE[(ri-3)%len(PALETTE)]
        ws2.cell(row=ri,column=1,value=cat).fill = PatternFill("solid",fgColor=bg+"22")
        ac = ws2.cell(row=ri,column=2,value=amt)
        ac.number_format='#,##0.00'; ac.font=Font(bold=True,name='Calibri')
        ws2.cell(row=ri,column=3,value=f"{amt/gt*100:.1f}%")
        for ci in range(1,4): ws2.cell(row=ri,column=ci).border=_border()

    pie = PieChart(); pie.title="By Category"; pie.style=10
    pie.add_data(Reference(ws2,min_col=2,min_row=2,max_row=2+len(stats['category_breakdown'])),titles_from_data=True)
    pie.set_categories(Reference(ws2,min_col=1,min_row=3,max_row=2+len(stats['category_breakdown'])))
    pie.width,pie.height=17,13; ws2.add_chart(pie,"E3")

    # ── Sheet 3: Vendor ────────────────────────────────────
    ws3 = wb.create_sheet("By Vendor")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:B1')
    _h(ws3['A1'],bg="0F172A"); ws3['A1'].value="Vendor Summary"
    for ci,h in enumerate(["Vendor","Total"],1):
        c=ws3.cell(row=2,column=ci,value=h); _h(c,bg="1E293B")
        ws3.column_dimensions[get_column_letter(ci)].width=[30,18][ci-1]
    for ri,(v,amt) in enumerate(stats['vendor_breakdown'].items(),3):
        ws3.cell(row=ri,column=1,value=v).border=_border()
        ac=ws3.cell(row=ri,column=2,value=amt)
        ac.number_format='#,##0.00'; ac.font=Font(bold=True,name='Calibri'); ac.border=_border()
    bar=BarChart(); bar.title="Vendor Spend"; bar.style=10; bar.type="bar"
    bar.add_data(Reference(ws3,min_col=2,min_row=2,max_row=2+len(stats['vendor_breakdown'])),titles_from_data=True)
    bar.set_categories(Reference(ws3,min_col=1,min_row=3,max_row=2+len(stats['vendor_breakdown'])))
    bar.width,bar.height=20,13; ws3.add_chart(bar,"D3")

    # ── Sheet 4: Monthly ───────────────────────────────────
    ws4 = wb.create_sheet("Monthly")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells('A1:B1')
    _h(ws4['A1'],bg="0F172A"); ws4['A1'].value="Monthly Summary"
    for ci,h in enumerate(["Month","Total"],1):
        c=ws4.cell(row=2,column=ci,value=h); _h(c,bg="1E293B")
        ws4.column_dimensions[get_column_letter(ci)].width=[16,18][ci-1]
    for ri,(month,amt) in enumerate(stats['monthly'].items(),3):
        ws4.cell(row=ri,column=1,value=month).border=_border()
        ac=ws4.cell(row=ri,column=2,value=amt)
        ac.number_format='#,##0.00'; ac.font=Font(bold=True,name='Calibri'); ac.border=_border()

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = proj_name.replace(' ','_').replace('/','_')
    path = os.path.join(EXPORT_DIR, f'{safe_name}_{ts}.xlsx')
    wb.save(path)
    return path
